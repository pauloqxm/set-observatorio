from __future__ import annotations

import io
import logging
import time
from typing import Any
from urllib.error import URLError
from urllib.request import urlopen

logger = logging.getLogger(__name__)

from openpyxl import load_workbook

WORKBOOK_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRg7NchDQ_1Sk7xbkyF3p8SDurWBGvU2WF_FOFCHgDLBsiqlZGlTHsUt-FMgun6hLCyPMbRO9HYkTU3/"
    "pub?output=xlsx"
)
CACHE_TTL_SECONDS = 300

VIRTUAL_SHEET_NAMES = frozenset({"mapa_regioes"})

_CACHE: dict[str, list[dict[str, Any]]] = {}
_SHEETS: list[str] = []
_LAST_FETCH_TS: float | None = None


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_number(raw_value: str) -> float:
    s = str(raw_value).strip().replace(" ", "")
    if not s:
        return 0.0
    negative = s.startswith("-")
    if negative:
        s = s[1:]
    if "," in s:
        # Formato BR típico: 1.234,56
        clean = s.replace(".", "").replace(",", ".")
    else:
        parts = s.split(".")
        if len(parts) == 2 and parts[1].isdigit() and len(parts[1]) <= 2:
            # Decimal internacional ou export Excel: 6629.0, 5.1
            clean = s
        else:
            # Só separadores de milhar: 1.234.567
            clean = s.replace(".", "")
    filtered = "".join(ch for ch in clean if ch.isdigit() or ch in ".-")
    try:
        v = float(filtered)
        return -v if negative else v
    except ValueError:
        return 0.0


def _load_workbook_cache() -> None:
    global _LAST_FETCH_TS, _CACHE, _SHEETS

    now = time.time()
    if _LAST_FETCH_TS and _CACHE and (now - _LAST_FETCH_TS) < CACHE_TTL_SECONDS:
        return

    try:
        raw_content = urlopen(WORKBOOK_URL, timeout=20).read()
    except (URLError, OSError) as exc:
        logger.error("Falha ao buscar planilha Google Sheets: %s", exc)
        if not _CACHE:
            raise RuntimeError(
                "Não foi possível carregar os dados da planilha. Verifique a conexão."
            ) from exc
        logger.warning("Usando cache expirado da planilha.")
        return

    try:
        wb = load_workbook(io.BytesIO(raw_content), read_only=True, data_only=True)
    except Exception as exc:
        logger.error("Falha ao processar arquivo xlsx da planilha: %s", exc)
        if not _CACHE:
            raise RuntimeError("Arquivo recebido do Google Sheets é inválido.") from exc
        logger.warning("Mantendo cache anterior após erro de leitura.")
        return

    data: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers_row = next(rows, None)
        if not headers_row:
            data[sheet_name] = []
            continue

        headers = [_normalize_cell(item) for item in headers_row]
        items: list[dict[str, Any]] = []
        for row in rows:
            record: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                record[header] = _normalize_cell(row[idx] if idx < len(row) else "")
            if any(str(v).strip() for v in record.values()):
                if "valor" in record:
                    record["valor_numerico"] = _to_number(record["valor"])
                items.append(record)
        data[sheet_name] = items

    _CACHE = data
    _SHEETS = wb.sheetnames
    _LAST_FETCH_TS = now


def get_sheet_names() -> list[str]:
    _load_workbook_cache()
    out = list(_SHEETS)
    for name in sorted(VIRTUAL_SHEET_NAMES):
        if name not in out:
            out.append(name)
    return out


def get_sheet_data(sheet_name: str) -> list[dict[str, Any]]:
    _load_workbook_cache()
    if sheet_name in VIRTUAL_SHEET_NAMES:
        return []
    return _CACHE.get(sheet_name, [])


def get_indicadores() -> list[dict[str, Any]]:
    return get_sheet_data("indicadores")
